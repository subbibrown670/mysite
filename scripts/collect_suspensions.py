#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup
from bs4 import XMLParsedAsHTMLWarning
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

NGA_PREF_INFO_URL = "https://www.nga.gr.jp/pref_info/"
USER_AGENT = "Mozilla/5.0 (compatible; suspension-collector/1.0)"
REQUEST_TIMEOUT = 10
warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
warnings.filterwarnings(
    "ignore",
    message="It looks like you're using an HTML parser to parse an XML document.",
)

DISCLOSURE_TEXT_HINTS = [
    "食品衛生法違反者等の公表",
    "不利益処分",
    "営業停止",
    "食中毒",
    "公表年月日",
]
LINK_TEXT_KEYWORDS = [
    "食品衛生",
    "違反者",
    "違反",
    "公表",
    "営業停止",
    "食中毒",
]
LINK_URL_KEYWORDS = [
    "shokuhin",
    "eisei",
    "ihan",
    "teishi",
    "food",
]

KNOWN_DISCLOSURE_URLS = {
    "東京都": "https://www.hokeniryo1.metro.tokyo.lg.jp/shokuhin/ihan/kouhyou.html",
}


@dataclass
class Source:
    prefecture: str
    root_url: str
    disclosure_url: str = ""


def normalize_text(text: str) -> str:
    return " ".join(text.replace("\xa0", " ").split())


def fetch_html(session: requests.Session, url: str, timeout: int = REQUEST_TIMEOUT) -> str:
    res = session.get(url, timeout=timeout)
    res.raise_for_status()
    encoding = res.apparent_encoding or res.encoding or "utf-8"
    return res.content.decode(encoding, errors="ignore")


def canonicalize_url(base_url: str, href: str) -> str:
    abs_url = urljoin(base_url, href)
    p = urlparse(abs_url)
    if p.scheme not in {"http", "https"}:
        return ""
    clean = p._replace(fragment="")
    return urlunparse(clean)


def is_same_host(url_a: str, url_b: str) -> bool:
    return urlparse(url_a).netloc == urlparse(url_b).netloc


def score_link(url: str, text: str) -> int:
    score = 0
    lt = text.lower()
    lu = url.lower()
    for kw in LINK_TEXT_KEYWORDS:
        if kw in text:
            score += 2
    for kw in ["サイトマップ", "保健", "健康", "くらし"]:
        if kw in text:
            score += 1
    for kw in LINK_URL_KEYWORDS:
        if kw in lu:
            score += 1
    if "公表" in text and "違反" in text:
        score += 2
    if "食品衛生法違反者等の公表" in text:
        score += 5
    if "pdf" in lu:
        score -= 2
    if len(lt) <= 2:
        score -= 1
    return score


def page_disclosure_score(page_text: str) -> int:
    score = 0
    for hint in DISCLOSURE_TEXT_HINTS:
        if hint in page_text:
            score += 1
    if "食品衛生法違反者等の公表" in page_text:
        score += 3
    if "営業停止" in page_text and ("公表" in page_text or "違反" in page_text):
        score += 2
    return score


def fetch_prefecture_roots(session: requests.Session) -> list[Source]:
    html = fetch_html(session, NGA_PREF_INFO_URL)
    soup = BeautifulSoup(html, "html.parser")
    sources: list[Source] = []
    seen: set[str] = set()

    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = normalize_text(a.get_text(" ", strip=True))
        if "公式ホームページ" not in text:
            continue
        if not ("pref." in href or "metro.tokyo" in href or "web.pref" in href):
            continue
        m = re.match(r"^(.+?[都道府県])公式ホームページ$", text)
        if not m:
            continue
        pref = m.group(1)
        url = canonicalize_url(NGA_PREF_INFO_URL, href)
        if not url or pref in seen:
            continue
        seen.add(pref)
        sources.append(Source(prefecture=pref, root_url=url))

    return sources


def load_sources_csv(path: Path) -> dict[str, Source]:
    if not path.exists():
        return {}
    out: dict[str, Source] = {}
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pref = (row.get("prefecture") or "").strip()
            if not pref:
                continue
            out[pref] = Source(
                prefecture=pref,
                root_url=(row.get("root_url") or "").strip(),
                disclosure_url=(row.get("disclosure_url") or "").strip(),
            )
    return out


def save_sources_csv(path: Path, sources: list[Source]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["prefecture", "root_url", "disclosure_url"])
        writer.writeheader()
        for s in sources:
            writer.writerow(
                {
                    "prefecture": s.prefecture,
                    "root_url": s.root_url,
                    "disclosure_url": s.disclosure_url,
                }
            )


def discover_from_sitemaps(session: requests.Session, root_url: str) -> str:
    host = urlparse(root_url).netloc
    sitemap_queue = [
        urljoin(root_url, "/sitemap.xml"),
        urljoin(root_url, "/sitemap_index.xml"),
    ]
    page_urls: set[str] = set()
    seen_sitemaps: set[str] = set()

    try:
        robots = fetch_html(session, urljoin(root_url, "/robots.txt"), timeout=6)
        for line in robots.splitlines():
            if line.lower().startswith("sitemap:"):
                u = normalize_text(line.split(":", 1)[1])
                if u:
                    sitemap_queue.append(u)
    except Exception:
        pass

    while sitemap_queue and len(seen_sitemaps) < 8:
        sm = sitemap_queue.pop(0)
        if sm in seen_sitemaps:
            continue
        seen_sitemaps.add(sm)
        try:
            body = fetch_html(session, sm, timeout=10)
        except Exception:
            continue
        if "<loc>" not in body:
            continue
        locs = re.findall(r"<loc>(.*?)</loc>", body, flags=re.IGNORECASE)
        if "<sitemapindex" in body:
            for loc in locs[:30]:
                u = normalize_text(loc)
                if u and urlparse(u).netloc == host:
                    sitemap_queue.append(u)
            continue
        for loc in locs:
            u = normalize_text(loc)
            if u and urlparse(u).netloc == host:
                page_urls.add(u)

    if not page_urls:
        return ""

    scored = sorted(
        [(score_link(u, ""), u) for u in page_urls],
        key=lambda x: x[0],
        reverse=True,
    )
    best_url = ""
    best_score = -1
    for _, u in scored[:60]:
        try:
            html = fetch_html(session, u, timeout=8)
        except Exception:
            continue
        text = normalize_text(BeautifulSoup(html, "html.parser").get_text(" ", strip=True))
        s = page_disclosure_score(text)
        if s > best_score:
            best_score = s
            best_url = u
        if s >= 5:
            return u

    return best_url if best_score >= 3 else ""


def discover_disclosure_url(session: requests.Session, root_url: str) -> str:
    sitemap_best = discover_from_sitemaps(session, root_url)
    if sitemap_best:
        return sitemap_best

    visited: set[str] = set()
    queue: list[tuple[str, int]] = [
        (root_url, 0),
        (urljoin(root_url, "/sitemap.xml"), 0),
        (urljoin(root_url, "/sitemap_index.xml"), 0),
    ]
    best_url = ""
    best_score = -1

    while queue and len(visited) < 12:
        current, depth = queue.pop(0)
        if current in visited:
            continue
        visited.add(current)

        try:
            html = fetch_html(session, current)
        except Exception:
            continue

        # sitemap XML から関連URL候補を抽出
        if "<urlset" in html or "<sitemapindex" in html:
            locs = re.findall(r"<loc>(.*?)</loc>", html, flags=re.IGNORECASE)
            xml_candidates: list[tuple[int, str]] = []
            for loc in locs:
                u = normalize_text(loc)
                if not u or not is_same_host(root_url, u):
                    continue
                s = score_link(u, "")
                if s <= 0:
                    continue
                xml_candidates.append((s, u))
            xml_candidates.sort(key=lambda x: x[0], reverse=True)
            for _, u in xml_candidates[:10]:
                if u not in visited:
                    queue.append((u, depth + 1))
            continue

        soup = BeautifulSoup(html, "html.parser")
        page_text = normalize_text(soup.get_text(" ", strip=True))
        pscore = page_disclosure_score(page_text)
        if pscore > best_score:
            best_score = pscore
            best_url = current

        if depth >= 2:
            continue

        link_candidates: list[tuple[int, str]] = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            text = normalize_text(a.get_text(" ", strip=True))
            abs_url = canonicalize_url(current, href)
            if not abs_url:
                continue
            if not is_same_host(root_url, abs_url):
                continue
            if abs_url in visited:
                continue
            lu = abs_url.lower()
            if any(lu.endswith(ext) for ext in [".jpg", ".jpeg", ".png", ".gif", ".zip", ".doc", ".docx", ".xls", ".xlsx"]):
                continue
            score = score_link(abs_url, text)
            if "sitemap" in lu or "サイトマップ" in text:
                score += 4
            if score <= 0 and depth > 0:
                continue
            link_candidates.append((score, abs_url))

        link_candidates.sort(key=lambda x: x[0], reverse=True)
        take = 10 if depth == 0 else 5
        for _, u in link_candidates[:take]:
            queue.append((u, depth + 1))

    if best_score >= 3:
        return best_url
    return ""


def split_facility_owner(block: str) -> tuple[str, str]:
    if not block:
        return "", ""
    facility = ""
    owner = ""
    m_fac = re.search(r"（施設の名称）\s*(.+?)(?=\s*（営業者氏名）|\s*（届出者氏名）|$)", block)
    if m_fac:
        facility = m_fac.group(1).strip()
    m_owner = re.search(r"（(?:営業者|届出者)氏名）\s*(.+)$", block)
    if m_owner:
        owner = m_owner.group(1).strip()
    return facility, owner


def parse_key_value_tables(soup: BeautifulSoup, prefecture: str, url: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for table in soup.find_all("table"):
        kv: dict[str, str] = {}
        for tr in table.find_all("tr"):
            tds = tr.find_all(["td", "th"])
            if len(tds) < 2:
                continue
            key = normalize_text(tds[0].get_text(" ", strip=True))
            val = normalize_text(tds[1].get_text(" ", strip=True))
            if key and val:
                kv[key] = val
        if "公表年月日" not in kv:
            continue
        merged = " ".join(kv.values())
        if "営業停止" not in merged:
            continue

        facility_owner = kv.get("施設の名称及び 営業者氏名等", "") or kv.get(
            "施設の名称及び 届出者氏名等", ""
        )
        facility, owner = split_facility_owner(facility_owner)

        rows.append(
            {
                "都道府県": prefecture,
                "公表年月日": kv.get("公表年月日", ""),
                "業種等": kv.get("業種等", ""),
                "施設名称": facility,
                "営業者氏名等": owner,
                "施設所在地等": kv.get("施設所在地等", ""),
                "不利益処分等の内容": kv.get("不利益処分等の内容", ""),
                "不利益処分等を行った理由": kv.get("不利益処分等を行った理由", ""),
                "備考": kv.get("備考", ""),
                "情報元URL": url,
                "抽出方式": "key_value",
            }
        )
    return rows


def parse_matrix_tables(soup: BeautifulSoup, prefecture: str, url: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for table in soup.find_all("table"):
        tr_list = table.find_all("tr")
        if len(tr_list) < 2:
            continue

        header_cells = tr_list[0].find_all(["th", "td"])
        headers = [normalize_text(c.get_text(" ", strip=True)) for c in header_cells]
        joined_header = " ".join(headers)
        if not any(k in joined_header for k in ["公表", "処分", "営業", "違反"]):
            continue

        col_date = next((i for i, h in enumerate(headers) if "公表" in h and "日" in h), -1)
        col_name = next((i for i, h in enumerate(headers) if "名称" in h or "施設" in h), -1)
        col_content = next((i for i, h in enumerate(headers) if "処分" in h or "内容" in h), -1)
        col_reason = next((i for i, h in enumerate(headers) if "理由" in h or "違反" in h), -1)
        col_addr = next((i for i, h in enumerate(headers) if "所在地" in h or "住所" in h), -1)
        col_type = next((i for i, h in enumerate(headers) if "業種" in h), -1)

        for tr in tr_list[1:]:
            cells = tr.find_all(["td", "th"])
            vals = [normalize_text(c.get_text(" ", strip=True)) for c in cells]
            if not vals:
                continue
            merged = " ".join(vals)
            if "営業停止" not in merged:
                continue

            def pick(idx: int) -> str:
                if idx < 0 or idx >= len(vals):
                    return ""
                return vals[idx]

            date_val = pick(col_date)
            content_val = pick(col_content) or merged
            if date_val in {"公表年月日", "不利益処分等の内容"}:
                continue
            if content_val.strip() in {"不利益処分等の内容", "営業停止"}:
                continue

            rows.append(
                {
                    "都道府県": prefecture,
                    "公表年月日": date_val,
                    "業種等": pick(col_type),
                    "施設名称": pick(col_name),
                    "営業者氏名等": "",
                    "施設所在地等": pick(col_addr),
                    "不利益処分等の内容": content_val,
                    "不利益処分等を行った理由": pick(col_reason),
                    "備考": "",
                    "情報元URL": url,
                    "抽出方式": "matrix",
                }
            )
    return rows


def parse_disclosure_page(html: str, prefecture: str, url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    rows = parse_key_value_tables(soup, prefecture, url)
    rows.extend(parse_matrix_tables(soup, prefecture, url))

    dedup: list[dict[str, str]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for r in rows:
        key = (
            r.get("都道府県", ""),
            r.get("公表年月日", ""),
            r.get("施設名称", ""),
            r.get("不利益処分等の内容", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        dedup.append(r)
    return dedup


def build_workbook(
    rows: list[dict[str, str]],
    output_path: Path,
    source_status: list[dict[str, str]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "営業停止情報"

    headers = [
        "更新日時",
        "都道府県",
        "公表年月日",
        "業種等",
        "施設名称",
        "営業者氏名等",
        "施設所在地等",
        "不利益処分等の内容",
        "不利益処分等を行った理由",
        "備考",
        "情報元URL",
        "抽出方式",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in rows:
        ws.append(
            [
                updated_at,
                row.get("都道府県", ""),
                row.get("公表年月日", ""),
                row.get("業種等", ""),
                row.get("施設名称", ""),
                row.get("営業者氏名等", ""),
                row.get("施設所在地等", ""),
                row.get("不利益処分等の内容", ""),
                row.get("不利益処分等を行った理由", ""),
                row.get("備考", ""),
                row.get("情報元URL", ""),
                row.get("抽出方式", ""),
            ]
        )

    widths = {
        1: 20,
        2: 10,
        3: 15,
        4: 16,
        5: 28,
        6: 35,
        7: 35,
        8: 30,
        9: 24,
        10: 36,
        11: 55,
        12: 12,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    src = wb.create_sheet("収集対象")
    src_headers = ["都道府県", "公式HP", "公表ページ", "収集件数", "状態"]
    src.append(src_headers)
    for c in range(1, len(src_headers) + 1):
        cell = src.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F6228")
    for s in source_status:
        src.append([s["都道府県"], s["公式HP"], s["公表ページ"], s["収集件数"], s["状態"]])
    src.column_dimensions["A"].width = 10
    src.column_dimensions["B"].width = 40
    src.column_dimensions["C"].width = 60
    src.column_dimensions["D"].width = 10
    src.column_dimensions["E"].width = 20

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def merge_sources(pref_roots: list[Source], cached: dict[str, Source]) -> list[Source]:
    merged: list[Source] = []
    for src in pref_roots:
        old = cached.get(src.prefecture)
        if old:
            root = old.root_url or src.root_url
            disclosure = KNOWN_DISCLOSURE_URLS.get(src.prefecture, old.disclosure_url)
            merged.append(Source(src.prefecture, root, disclosure))
        else:
            merged.append(
                Source(
                    src.prefecture,
                    src.root_url,
                    KNOWN_DISCLOSURE_URLS.get(src.prefecture, ""),
                )
            )
    return merged


def main() -> int:
    parser = argparse.ArgumentParser(
        description="都道府県の飲食店営業停止情報を収集し、Excelを更新します。"
    )
    parser.add_argument(
        "--output",
        default="output/spreadsheet/eigyo_teishi_summary.xlsx",
        help="出力先Excelパス",
    )
    parser.add_argument(
        "--sources-csv",
        default="data/prefecture_sources.csv",
        help="都道府県ごとの公表ページキャッシュCSV",
    )
    parser.add_argument(
        "--refresh-sources",
        action="store_true",
        help="公表ページURLを全都道府県で再探索する",
    )
    args = parser.parse_args()

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    roots = fetch_prefecture_roots(session)
    cached = load_sources_csv(Path(args.sources_csv))
    sources = merge_sources(roots, cached)

    all_rows: list[dict[str, str]] = []
    source_status: list[dict[str, str]] = []

    for src in sources:
        known_url = KNOWN_DISCLOSURE_URLS.get(src.prefecture, "")
        disclosure_url = known_url or src.disclosure_url
        if not known_url and (args.refresh_sources or not disclosure_url):
            disclosure_url = discover_disclosure_url(session, src.root_url)
            src.disclosure_url = disclosure_url

        rows: list[dict[str, str]] = []
        status = "未検出"
        if disclosure_url:
            try:
                html = fetch_html(session, disclosure_url)
                rows = parse_disclosure_page(html, src.prefecture, disclosure_url)
                status = "取得済み" if rows else "ページ発見(0件)"
            except Exception:
                status = "取得失敗"

        all_rows.extend(rows)
        source_status.append(
            {
                "都道府県": src.prefecture,
                "公式HP": src.root_url,
                "公表ページ": disclosure_url,
                "収集件数": str(len(rows)),
                "状態": status,
            }
        )
        print(f"{src.prefecture}: {status} rows={len(rows)} url={disclosure_url}", flush=True)

    save_sources_csv(Path(args.sources_csv), sources)
    build_workbook(all_rows, Path(args.output), source_status)

    print(f"updated: {args.output} total_rows={len(all_rows)} prefectures={len(sources)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
