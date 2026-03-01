#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

USER_AGENT = "Mozilla/5.0 (compatible; food-poisoning-collector/1.1)"
REQUEST_TIMEOUT = 12

HEADERS = [
    "更新日時",
    "都道府県",
    "保健所名",
    "発表機関",
    "参照元URL",
    "最終更新日(掲載日)",
    "事案名/件名",
    "発生日",
    "原因施設/施設名",
    "所在地",
    "原因食品",
    "病因物質",
    "患者数",
    "死者数",
    "概要",
    "対応状況",
    "取得ステータス",
    "リトライ回数",
    "取得エラー分類",
    "取得エラー詳細",
]

MISSING = "ー"

DATE_RE = re.compile(r"(20\d{2}[./年-]\d{1,2}[./月-]\d{1,2}日?)")
PATIENT_RE = re.compile(r"(\d+)\s*名")
DEATH_RE = re.compile(r"(\d+)\s*名")
FOOD_RE = re.compile(r"(?:原因食品|食品)[:：]?\s*([^\s、。]{2,40})")
PATHOGEN_RE = re.compile(r"(?:病因物質|原因物質)[:：]?\s*([^\s、。]{2,40})")
FACILITY_RE = re.compile(r"(?:施設名|原因施設|店舗名|店名)[:：]?\s*([^\n]{2,60})")
ADDRESS_RE = re.compile(r"(?:所在地|住所)[:：]?\s*([^\n]{2,80})")
HOKENJO_RE = re.compile(r"(?:保健所名|所管保健所|保健所)[:：]?\s*([^\n]{2,60})")
ORG_RE = re.compile(r"(?:発表機関|公表機関|自治体|都道府県)[:：]?\s*([^\n]{2,60})")

PREF_ALIASES = {"prefecture", "都道府県", "都道府県名", "pref", "name"}
ROOT_ALIASES = {"root_url", "公式hp", "公式url", "root", "url", "hp"}
DISC_ALIASES = {
    "disclosure_url",
    "公表ページ",
    "参照url",
    "参照先url",
    "食中毒url",
    "food_poisoning_url",
    "source_url",
}


@dataclass
class Source:
    prefecture: str
    root_url: str
    disclosure_url: str


class FetchAttemptError(Exception):
    def __init__(self, cause: Exception, retries_used: int) -> None:
        super().__init__(str(cause))
        self.cause = cause
        self.retries_used = retries_used


def norm(text: str) -> str:
    if text is None:
        return ""
    return " ".join(str(text).replace("\xa0", " ").split())


def normalize_key(key: str) -> str:
    return norm(key).lower().replace(" ", "").replace("_", "")


def pick_value_from_record(record: dict[str, str], aliases: set[str]) -> str:
    normalized_aliases = {normalize_key(a) for a in aliases}
    for k, v in record.items():
        if normalize_key(k) in normalized_aliases:
            return norm(v)
    return ""


def load_sources_csv(path: Path) -> list[Source]:
    out: list[Source] = []
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pref = pick_value_from_record(row, PREF_ALIASES)
            if not pref:
                continue
            root = pick_value_from_record(row, ROOT_ALIASES)
            disclosure = pick_value_from_record(row, DISC_ALIASES)
            out.append(Source(pref, root, disclosure or root))
    return out


def load_sources_xlsx(path: Path, sheet_name: str = "") -> list[Source]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []
    headers = [norm(v) for v in values[0]]
    out: list[Source] = []
    for row_vals in values[1:]:
        record = {headers[i]: norm(row_vals[i]) if i < len(row_vals) else "" for i in range(len(headers))}
        pref = pick_value_from_record(record, PREF_ALIASES)
        if not pref:
            continue
        root = pick_value_from_record(record, ROOT_ALIASES)
        disclosure = pick_value_from_record(record, DISC_ALIASES)
        out.append(Source(pref, root, disclosure or root))
    return out


def load_sources(path: Path, sheet_name: str = "") -> list[Source]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_sources_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_sources_xlsx(path, sheet_name=sheet_name)
    raise ValueError(f"unsupported source file type: {suffix}")


def is_retryable_error(exc: Exception) -> bool:
    if isinstance(exc, (requests.exceptions.Timeout, requests.exceptions.ConnectionError)):
        return True
    if isinstance(exc, requests.exceptions.HTTPError):
        status = exc.response.status_code if exc.response is not None else 0
        return status == 429 or 500 <= status <= 599
    return False


def classify_error(exc: Exception) -> str:
    if isinstance(exc, requests.exceptions.Timeout):
        return "タイムアウト"
    if isinstance(exc, requests.exceptions.SSLError):
        return "SSLエラー"
    if isinstance(exc, requests.exceptions.HTTPError):
        status = exc.response.status_code if exc.response is not None else 0
        return f"HTTP_{status}" if status else "HTTPエラー"
    msg = str(exc).lower()
    if "name resolution" in msg or "failed to resolve" in msg or "nodename nor servname" in msg:
        return "DNS失敗"
    if isinstance(exc, requests.exceptions.ConnectionError):
        return "接続失敗"
    return "不明エラー"


def fetch_html_with_retry(
    session: requests.Session,
    url: str,
    timeout: int,
    max_retries: int,
    backoff_sec: float,
) -> tuple[str, int]:
    retries_used = 0
    for attempt in range(max_retries + 1):
        try:
            res = session.get(url, timeout=timeout)
            res.raise_for_status()
            enc = res.apparent_encoding or res.encoding or "utf-8"
            return res.content.decode(enc, errors="ignore"), retries_used
        except Exception as e:
            if attempt >= max_retries or not is_retryable_error(e):
                raise FetchAttemptError(e, retries_used) from e
            retries_used += 1
            time.sleep(backoff_sec * (2**attempt))
    raise FetchAttemptError(RuntimeError("unreachable"), retries_used)


def pick_first(pattern: re.Pattern[str], text: str) -> str:
    m = pattern.search(text)
    return m.group(1).strip() if m else MISSING


def extract_blocks(soup: BeautifulSoup) -> Iterable[str]:
    blocks: list[str] = []
    for tag in soup.find_all(["tr", "li", "p", "article", "section", "div"]):
        text = norm(tag.get_text(" ", strip=True))
        if len(text) < 16:
            continue
        if "食中毒" in text:
            blocks.append(text)
    if not blocks:
        whole = norm(soup.get_text(" ", strip=True))
        if "食中毒" in whole:
            blocks.append(whole)
    return blocks[:20]


def base_row(prefecture: str, url: str) -> dict[str, str]:
    return {
        "更新日時": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "都道府県": prefecture,
        "保健所名": MISSING,
        "発表機関": MISSING,
        "参照元URL": url or MISSING,
        "最終更新日(掲載日)": MISSING,
        "事案名/件名": MISSING,
        "発生日": MISSING,
        "原因施設/施設名": MISSING,
        "所在地": MISSING,
        "原因食品": MISSING,
        "病因物質": MISSING,
        "患者数": MISSING,
        "死者数": MISSING,
        "概要": MISSING,
        "対応状況": MISSING,
        "取得ステータス": MISSING,
        "リトライ回数": "0",
        "取得エラー分類": MISSING,
        "取得エラー詳細": MISSING,
    }


def parse_latest_event(prefecture: str, url: str, html: str) -> dict[str, str]:
    row = base_row(prefecture, url)
    soup = BeautifulSoup(html, "html.parser")
    blocks = list(extract_blocks(soup))

    if not blocks:
        row["取得ステータス"] = "食中毒情報なし"
        return row

    block = blocks[0]
    date_values = DATE_RE.findall(block)
    posting = date_values[0] if len(date_values) >= 1 else MISSING
    onset = date_values[1] if len(date_values) >= 2 else MISSING
    p_m = PATIENT_RE.search(block)
    d_m = DEATH_RE.search(block) if "死亡" in block else None

    row.update(
        {
            "保健所名": pick_first(HOKENJO_RE, block),
            "発表機関": pick_first(ORG_RE, block),
            "最終更新日(掲載日)": posting,
            "事案名/件名": "食中毒事案" if "食中毒" in block else MISSING,
            "発生日": onset,
            "原因施設/施設名": pick_first(FACILITY_RE, block),
            "所在地": pick_first(ADDRESS_RE, block),
            "原因食品": pick_first(FOOD_RE, block),
            "病因物質": pick_first(PATHOGEN_RE, block),
            "患者数": p_m.group(1) if p_m else MISSING,
            "死者数": d_m.group(1) if d_m else MISSING,
            "概要": block[:220],
            "対応状況": "公表中",
            "取得ステータス": "取得済み",
        }
    )
    return row


def build_workbook(rows: list[dict[str, str]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "食中毒情報"
    ws.append(HEADERS)

    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for row in rows:
        ws.append([row.get(h, MISSING) or MISSING for h in HEADERS])

    widths = {
        1: 20,
        2: 10,
        3: 16,
        4: 16,
        5: 48,
        6: 16,
        7: 18,
        8: 16,
        9: 24,
        10: 24,
        11: 16,
        12: 16,
        13: 10,
        14: 10,
        15: 45,
        16: 14,
        17: 14,
        18: 10,
        19: 14,
        20: 55,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="都道府県の食中毒情報を収集してExcelを更新")
    parser.add_argument(
        "--sources",
        "--sources-csv",
        dest="sources",
        default="data/prefecture_sources.csv",
        help="入力ファイル (.csv または .xlsx)",
    )
    parser.add_argument(
        "--sources-sheet",
        default="",
        help="入力が .xlsx の場合のシート名（省略時は先頭シート）",
    )
    parser.add_argument("--output", default="output/spreadsheet/food_poisoning_summary.xlsx")
    parser.add_argument("--timeout", type=int, default=REQUEST_TIMEOUT)
    parser.add_argument("--max-retries", type=int, default=2)
    parser.add_argument("--retry-backoff", type=float, default=1.0)
    args = parser.parse_args()

    sources = load_sources(Path(args.sources), sheet_name=args.sources_sheet)
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    rows: list[dict[str, str]] = []
    for src in sources:
        url = src.disclosure_url or src.root_url
        if not url:
            row = base_row(src.prefecture, "")
            row["取得ステータス"] = "取得失敗"
            row["取得エラー分類"] = "入力不備"
            row["取得エラー詳細"] = "URLが空です"
            rows.append(row)
            continue

        try:
            html, retries = fetch_html_with_retry(
                session=session,
                url=url,
                timeout=args.timeout,
                max_retries=args.max_retries,
                backoff_sec=args.retry_backoff,
            )
            row = parse_latest_event(src.prefecture, url, html)
            row["リトライ回数"] = str(retries)
        except FetchAttemptError as e:
            row = base_row(src.prefecture, url)
            row["取得ステータス"] = "取得失敗"
            row["リトライ回数"] = str(e.retries_used)
            row["取得エラー分類"] = classify_error(e.cause)
            row["取得エラー詳細"] = str(e.cause) or MISSING
        except Exception as e:
            row = base_row(src.prefecture, url)
            row["取得ステータス"] = "取得失敗"
            row["取得エラー分類"] = "不明エラー"
            row["取得エラー詳細"] = str(e) or MISSING

        rows.append(row)

    build_workbook(rows, Path(args.output))
    print(f"updated: {args.output} rows={len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
